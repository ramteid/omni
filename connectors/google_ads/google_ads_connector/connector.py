"""Main Google Ads connector."""

from __future__ import annotations

import csv
import io
import logging
from datetime import UTC, datetime, timedelta
from typing import Any, TypedDict, cast

from fastapi.responses import JSONResponse, Response
from omni_connector import Connector, SearchOperator, SyncContext, SyncMode
from omni_connector.models import (
    ActionDefinition,
    ActionResponse,
    OAuthManifestConfig,
    OAuthScopeSet,
)

from .client import (
    GoogleAdsApiError,
    GoogleAdsClient,
    GoogleAdsConnectorError,
    InMemoryGoogleAdsClient,
)
from .config import GOOGLE_ADS_SCOPE, GoogleAdsCredentials, GoogleAdsSourceConfig
from .mappers import map_row_to_document, render_content
from .models import (
    CHANGE_STATUS_QUERY_TEMPLATE,
    REPORT_RESOURCE_ALLOWLIST,
    SYNC_QUERIES,
)

logger = logging.getLogger(__name__)

DEFAULT_REPORT_LIMIT = 5000
DEFAULT_EXPORT_LIMIT = 10000
MAX_JSON_ROWS = 10000
MAX_CSV_ROWS = 25000
MAX_XLSX_ROWS = 10000
CHECKPOINT_SCHEMA_VERSION = 1

METRIC_FIELDS = {
    "metrics.impressions",
    "metrics.clicks",
    "metrics.cost_micros",
    "metrics.ctr",
    "metrics.average_cpc",
    "metrics.conversions",
    "metrics.conversions_value",
    "metrics.all_conversions",
    "metrics.cost_per_conversion",
    "metrics.search_impression_share",
    "metrics.search_top_impression_share",
    "metrics.search_absolute_top_impression_share",
}

SEGMENT_FIELDS = {
    "segments.date",
    "segments.week",
    "segments.month",
    "segments.device",
    "segments.ad_network_type",
    "segments.day_of_week",
    "segments.hour",
    "segments.conversion_action",
    "segments.conversion_action_name",
}

SEGMENTED_REPORT_RESOURCES = {
    "customer",
    "campaign",
    "ad_group",
    "ad_group_ad",
    "keyword_view",
}

REPORT_BUILDERS = {
    "campaign_performance": "campaign",
    "ad_group_performance": "ad_group",
    "keyword_performance": "keyword_view",
    "search_term": "search_term_view",
    "ad_creative": "ad_group_ad",
    "asset_performance": "ad_group_ad_asset_view",
    "landing_page": "expanded_landing_page_view",
    "budget_pacing": "campaign_budget",
    "conversion_performance": "campaign",
    "segmented_performance": "segmented",
    "change_history": "change_event",
    "policy_diagnostics": "ad_group_ad",
}


class GoogleAdsCheckpoint(TypedDict):
    schema_version: int
    last_successful_sync_at: str | None
    last_completed_unit: str | None


# Reserved for future non-sync source-level metadata.
class GoogleAdsConnectorState(TypedDict):
    pass


def _checkpoint(
    *,
    last_successful_sync_at: str | None,
    last_completed_unit: str | None,
) -> GoogleAdsCheckpoint:
    return {
        "schema_version": CHECKPOINT_SCHEMA_VERSION,
        "last_successful_sync_at": last_successful_sync_at,
        "last_completed_unit": last_completed_unit,
    }


def _parse_checkpoint(raw: dict[str, Any] | None) -> GoogleAdsCheckpoint:
    if not raw:
        return _checkpoint(last_successful_sync_at=None, last_completed_unit=None)

    if raw.get("schema_version") != CHECKPOINT_SCHEMA_VERSION:
        raise ValueError("Invalid Google Ads checkpoint schema_version")

    last_successful_sync_at = raw.get("last_successful_sync_at")
    if last_successful_sync_at is not None and not isinstance(
        last_successful_sync_at, str
    ):
        raise ValueError("Invalid Google Ads checkpoint last_successful_sync_at")

    last_completed_unit = raw.get("last_completed_unit")
    if last_completed_unit is not None and not isinstance(last_completed_unit, str):
        raise ValueError("Invalid Google Ads checkpoint last_completed_unit")

    return _checkpoint(
        last_successful_sync_at=last_successful_sync_at,
        last_completed_unit=last_completed_unit,
    )


def _last_successful_sync_time(checkpoint: GoogleAdsCheckpoint) -> str | None:
    return checkpoint["last_successful_sync_at"]


def _last_completed_unit(checkpoint: GoogleAdsCheckpoint) -> str | None:
    return checkpoint["last_completed_unit"]


def _unit_key(customer_id: str, entity_type: str) -> str:
    return f"{customer_id}:{entity_type}"


class GoogleAdsConnector(Connector):
    """Google Ads connector for Omni."""

    @property
    def name(self) -> str:
        return "google_ads"

    @property
    def display_name(self) -> str:
        return "Google Ads"

    @property
    def version(self) -> str:
        return "1.0.0"

    @property
    def source_types(self) -> list[str]:
        return ["google_ads"]

    @property
    def description(self) -> str:
        return (
            "Index Google Ads account structure and run live campaign analysis reports"
        )

    @property
    def sync_modes(self) -> list[str]:
        return ["full", "incremental"]

    @property
    def search_operators(self) -> list[SearchOperator]:
        return [
            SearchOperator(
                operator="customer", attribute_key="customer_id", value_type="text"
            ),
            SearchOperator(
                operator="campaign", attribute_key="campaign_id", value_type="text"
            ),
            SearchOperator(
                operator="ad_group", attribute_key="ad_group_id", value_type="text"
            ),
            SearchOperator(operator="asset", attribute_key="asset_id", value_type="text"),
            SearchOperator(
                operator="criterion", attribute_key="criterion_id", value_type="text"
            ),
            SearchOperator(
                operator="shared_set", attribute_key="shared_set_id", value_type="text"
            ),
            SearchOperator(
                operator="status", attribute_key="status", value_type="text"
            ),
            SearchOperator(
                operator="channel", attribute_key="channel_type", value_type="text"
            ),
            SearchOperator(
                operator="entity", attribute_key="entity_type", value_type="text"
            ),
            SearchOperator(operator="label", attribute_key="labels", value_type="text"),
        ]

    def oauth_config(self) -> OAuthManifestConfig | None:
        return OAuthManifestConfig(
            provider="google_ads",
            auth_endpoint="https://accounts.google.com/o/oauth2/v2/auth",
            token_endpoint="https://oauth2.googleapis.com/token",
            userinfo_endpoint="https://www.googleapis.com/oauth2/v3/userinfo",
            userinfo_email_field="email",
            identity_scopes=["email", "profile"],
            scopes={
                "google_ads": OAuthScopeSet(
                    read=[GOOGLE_ADS_SCOPE],
                    write=[GOOGLE_ADS_SCOPE],
                )
            },
            extra_auth_params={"access_type": "offline", "prompt": "consent"},
            scope_separator=" ",
        )

    @property
    def actions(self) -> list[ActionDefinition]:
        report_actions = [
            _report_action_definition(
                "get_campaign_performance_report",
                "Fetch live campaign performance metrics by date/campaign.",
            ),
            _report_action_definition(
                "get_ad_group_performance_report",
                "Fetch live ad group performance metrics by date/campaign/ad group.",
            ),
            _report_action_definition(
                "get_keyword_performance_report",
                "Fetch live keyword performance metrics by date/campaign/ad group/keyword.",
            ),
            _report_action_definition(
                "get_search_term_report",
                "Fetch live search term performance for mining queries and waste.",
            ),
            _report_action_definition(
                "get_ad_creative_report",
                "Fetch live ad creative performance and ad text/status fields.",
            ),
            _report_action_definition(
                "get_asset_performance_report",
                "Fetch live asset performance labels and metrics where available.",
            ),
            _report_action_definition(
                "get_landing_page_report",
                "Fetch live expanded landing page performance.",
            ),
            _report_action_definition(
                "get_budget_pacing_report",
                "Fetch live budget, spend, and pacing data by campaign budget.",
            ),
            _report_action_definition(
                "get_conversion_performance_report",
                "Fetch live conversion performance by campaign and conversion action.",
            ),
            _report_action_definition(
                "get_segmented_performance_report",
                "Fetch live performance for a supported resource with selected segments.",
                extra_properties={
                    "resource": {
                        "type": "string",
                        "enum": sorted(SEGMENTED_REPORT_RESOURCES),
                        "default": "campaign",
                    },
                    "dimensions": {
                        "type": "array",
                        "items": {"type": "string", "enum": sorted(SEGMENT_FIELDS)},
                    },
                    "metrics": {
                        "type": "array",
                        "items": {"type": "string", "enum": sorted(METRIC_FIELDS)},
                    },
                },
            ),
        ]
        return [
            ActionDefinition(
                name="run_gaql_query",
                description=(
                    "Run a live Google Ads GAQL query for analysis. Returns structured JSON rows."
                ),
                mode="read",
                source_types=["google_ads"],
                input_schema={
                    "type": "object",
                    "properties": {
                        "customer_id": {"type": "string"},
                        "query": {"type": "string"},
                        "limit": {"type": "integer", "default": 1000, "maximum": 10000},
                    },
                    "required": ["customer_id", "query"],
                },
            ),
            ActionDefinition(
                name="export_gaql_report_csv",
                description="Run a live Google Ads GAQL query and return a CSV export.",
                mode="read",
                source_types=["google_ads"],
                input_schema=_gaql_export_schema(),
            ),
            ActionDefinition(
                name="export_gaql_report_xlsx",
                description="Run a live Google Ads GAQL query and return an XLSX export.",
                mode="read",
                source_types=["google_ads"],
                input_schema=_gaql_export_schema(),
            ),
            ActionDefinition(
                name="export_report",
                description=(
                    "Run a curated Google Ads report and export JSON, CSV, or XLSX with metadata."
                ),
                mode="read",
                source_types=["google_ads"],
                input_schema={
                    "type": "object",
                    "properties": {
                        "customer_id": {"type": "string"},
                        "report_type": {
                            "type": "string",
                            "enum": sorted(REPORT_BUILDERS),
                        },
                        "format": {
                            "type": "string",
                            "enum": ["json", "csv", "xlsx"],
                            "default": "csv",
                        },
                        **_report_common_schema_properties(),
                    },
                    "required": ["customer_id", "report_type"],
                },
            ),
            ActionDefinition(
                name="get_account_summary",
                description=(
                    "Fetch live customer/campaign structure and recent performance summary rows."
                ),
                mode="read",
                source_types=["google_ads"],
                input_schema={
                    "type": "object",
                    "properties": {
                        "customer_id": {"type": "string"},
                        "date_range": {"type": "string", "default": "LAST_30_DAYS"},
                    },
                    "required": ["customer_id"],
                },
            ),
            ActionDefinition(
                name="get_recommendations",
                description="Fetch live Google Ads recommendations for a customer.",
                mode="read",
                source_types=["google_ads"],
                input_schema={
                    "type": "object",
                    "properties": {"customer_id": {"type": "string"}},
                    "required": ["customer_id"],
                },
            ),
            ActionDefinition(
                name="get_change_history",
                description="Fetch recent Google Ads change history for root-cause analysis.",
                mode="read",
                source_types=["google_ads"],
                input_schema=_report_action_schema(),
            ),
            ActionDefinition(
                name="get_policy_diagnostics",
                description="Fetch live ad policy/approval diagnostics.",
                mode="read",
                source_types=["google_ads"],
                input_schema=_report_action_schema(),
            ),
            ActionDefinition(
                name="get_account_hierarchy",
                description="List accessible Google Ads accounts and customer-client hierarchy.",
                mode="read",
                source_types=["google_ads"],
                input_schema={
                    "type": "object",
                    "properties": {
                        "customer_id": {
                            "type": "string",
                            "description": (
                                "Optional manager/customer ID to query customer_client hierarchy."
                            ),
                        },
                        "limit": {"type": "integer", "default": 1000, "maximum": 10000},
                    },
                },
            ),
            *report_actions,
        ]

    async def sync(
        self,
        source_config: dict[str, Any],
        credentials: dict[str, Any],
        state: dict[str, Any] | None,
        ctx: SyncContext,
    ) -> None:
        try:
            raw_creds = credentials.get("credentials", credentials)
            merged_creds = (
                {**credentials, **raw_creds}
                if isinstance(raw_creds, dict)
                else credentials
            )
            creds = GoogleAdsCredentials.parse(merged_creds)
            config = GoogleAdsSourceConfig.parse(source_config, merged_creds)
            checkpoint = _parse_checkpoint(state)
        except ValueError as exc:
            await ctx.fail(str(exc))
            return

        if not config.sync_enabled:
            await ctx.complete(
                new_state=_checkpoint(
                    last_successful_sync_at=_last_successful_sync_time(checkpoint),
                    last_completed_unit=None,
                )
            )
            return

        client = self._make_client(creds, config, source_config)

        try:
            if ctx.sync_mode == SyncMode.INCREMENTAL:
                await self._incremental_sync(client, config, checkpoint, ctx)
            else:
                await self._full_sync(client, config, checkpoint, ctx)
        except GoogleAdsConnectorError as exc:
            logger.exception("Google Ads sync failed")
            await ctx.fail(str(exc))
        except Exception as exc:
            logger.exception("Unexpected Google Ads sync failure")
            await ctx.fail(str(exc))

    def _make_client(
        self,
        creds: GoogleAdsCredentials,
        config: GoogleAdsSourceConfig,
        source_config: dict[str, Any],
    ) -> GoogleAdsClient:
        if isinstance(source_config.get("mock_data"), dict):
            return InMemoryGoogleAdsClient(source_config["mock_data"])
        return GoogleAdsClient(creds, login_customer_id=config.login_customer_id)

    async def _full_sync(
        self,
        client: GoogleAdsClient,
        config: GoogleAdsSourceConfig,
        checkpoint: GoogleAdsCheckpoint,
        ctx: SyncContext,
    ) -> None:
        run_started_at = datetime.now(UTC).isoformat()
        previous_success = _last_successful_sync_time(checkpoint)
        units = [
            (customer_id, entity_type)
            for customer_id in sorted(config.customer_ids)
            for entity_type in config.entity_types
        ]
        last_completed_unit = _last_completed_unit(checkpoint) if ctx.is_resume else None
        skip_completed = last_completed_unit in {
            _unit_key(customer_id, entity_type) for customer_id, entity_type in units
        }

        for customer_id, entity_type in units:
            key = _unit_key(customer_id, entity_type)
            if skip_completed:
                if key == last_completed_unit:
                    skip_completed = False
                continue
            if ctx.is_cancelled():
                await ctx.fail("Cancelled by user")
                return
            completed_entity = await self._sync_entity_type(
                client, customer_id, entity_type, ctx
            )
            if not completed_entity:
                await ctx.fail("Cancelled by user")
                return
            await ctx.save_checkpoint(
                _checkpoint(
                    last_successful_sync_at=previous_success,
                    last_completed_unit=key,
                )
            )

        await ctx.complete(
            new_state=_checkpoint(
                last_successful_sync_at=run_started_at,
                last_completed_unit=None,
            )
        )

    async def _sync_entity_type(
        self,
        client: GoogleAdsClient,
        customer_id: str,
        entity_type: str,
        ctx: SyncContext,
    ) -> bool:
        query = SYNC_QUERIES.get(entity_type)
        if not query:
            return True
        logger.info("Syncing Google Ads %s for customer %s", entity_type, customer_id)
        try:
            async for row in client.search_stream(customer_id, query):
                if ctx.is_cancelled():
                    return False
                await ctx.increment_scanned()
                try:
                    content = render_content(entity_type, customer_id, row)
                    content_id = await ctx.content_storage.save(content, "text/plain")
                    doc = map_row_to_document(
                        entity_type=entity_type,
                        customer_id=customer_id,
                        row=row,
                        content_id=content_id,
                    )
                    await ctx.emit(doc)
                except Exception as exc:
                    logger.warning("Failed to map Google Ads row: %s", exc)
                    await ctx.emit_error(
                        f"google_ads:{customer_id}:{entity_type}:unknown", str(exc)
                    )
        except GoogleAdsApiError as exc:
            await ctx.emit_error(f"google_ads:{customer_id}:{entity_type}:*", str(exc))
        return True

    async def _incremental_sync(
        self,
        client: GoogleAdsClient,
        config: GoogleAdsSourceConfig,
        checkpoint: GoogleAdsCheckpoint,
        ctx: SyncContext,
    ) -> None:
        since = _last_successful_sync_time(checkpoint)
        if not since:
            await self._full_sync(client, config, checkpoint, ctx)
            return

        run_started_at = datetime.now(UTC).isoformat()
        units = [
            (customer_id, entity_type)
            for customer_id in sorted(config.customer_ids)
            for entity_type in config.entity_types
        ]
        last_completed_unit = _last_completed_unit(checkpoint) if ctx.is_resume else None
        skip_completed = last_completed_unit in {
            _unit_key(customer_id, entity_type) for customer_id, entity_type in units
        }
        changed_by_customer: dict[str, set[str]] = {}

        for customer_id, entity_type in units:
            key = _unit_key(customer_id, entity_type)
            if skip_completed:
                if key == last_completed_unit:
                    skip_completed = False
                continue
            if ctx.is_cancelled():
                await ctx.fail("Cancelled by user")
                return
            if customer_id not in changed_by_customer:
                changed_by_customer[customer_id] = await self._changed_resource_types(
                    client, customer_id, since, ctx
                )
            if entity_type not in changed_by_customer[customer_id]:
                continue
            completed_entity = await self._sync_entity_type(
                client, customer_id, entity_type, ctx
            )
            if not completed_entity:
                await ctx.fail("Cancelled by user")
                return
            await ctx.save_checkpoint(
                _checkpoint(
                    last_successful_sync_at=since,
                    last_completed_unit=key,
                )
            )

        await ctx.complete(
            new_state=_checkpoint(
                last_successful_sync_at=run_started_at,
                last_completed_unit=None,
            )
        )

    async def _changed_resource_types(
        self,
        client: GoogleAdsClient,
        customer_id: str,
        since: str,
        ctx: SyncContext,
    ) -> set[str]:
        query = CHANGE_STATUS_QUERY_TEMPLATE.format(since=since.replace("'", ""))
        resource_map = {
            "campaign": "campaign",
            "campaign_budget": "campaign_budget",
            "bidding_strategy": "bidding_strategy",
            "ad_group": "ad_group",
            "ad_group_ad": "ad_group_ad",
            "asset": "asset",
            "ad_group_criterion": "keyword_view",
            "user_list": "user_list",
            "conversion_action": "conversion_action",
            "shared_set": "shared_set",
        }
        changed: set[str] = set()
        try:
            async for row in client.search(customer_id, query):
                await ctx.increment_scanned()
                change_status = (
                    row.get("change_status") or row.get("changeStatus") or row
                )
                resource_type = str(
                    change_status.get("resource_type")
                    or change_status.get("resourceType")
                    or ""
                ).lower()
                if resource_type in resource_map:
                    changed.add(resource_map[resource_type])
        except GoogleAdsApiError as exc:
            await ctx.emit_error(f"google_ads:{customer_id}:change_status", str(exc))
        return changed

    async def execute_action(  # type: ignore[override]
        self,
        action: str,
        params: dict[str, Any],
        credentials: dict[str, Any],
    ) -> JSONResponse | Response:
        try:
            raw_creds = credentials.get("credentials", credentials)
            merged_creds = (
                {**credentials, **raw_creds}
                if isinstance(raw_creds, dict)
                else credentials
            )
            creds = GoogleAdsCredentials.parse(merged_creds)
            raw_source_config = params.get("source_config")
            source_config = cast(
                dict[str, Any],
                raw_source_config if isinstance(raw_source_config, dict) else {},
            )
            customer_id = _optional_customer_id(params)
            action_config = {**source_config}
            if customer_id:
                action_config["customer_ids"] = [customer_id]
            elif action == "get_account_hierarchy" and not _has_configured_customer_ids(
                action_config, merged_creds
            ):
                action_config["customer_ids"] = ["0"]
            cfg = GoogleAdsSourceConfig.parse(action_config, merged_creds)
            client = self._make_client(creds, cfg, source_config)
        except Exception as exc:
            return ActionResponse.failure(str(exc)).to_response(status_code=400)

        try:
            if action == "run_gaql_query":
                return await self._action_run_gaql(client, params)
            if action == "export_gaql_report_csv":
                return await self._action_export_csv(client, params)
            if action == "export_gaql_report_xlsx":
                return await self._action_export_xlsx(client, params)
            if action == "export_report":
                return await self._action_export_report(client, params)
            if action == "get_account_summary":
                return await self._action_account_summary(client, params)
            if action == "get_recommendations":
                return await self._action_recommendations(client, params)
            if action == "get_change_history":
                return await self._action_curated_report(client, params, "change_history")
            if action == "get_policy_diagnostics":
                return await self._action_curated_report(client, params, "policy_diagnostics")
            if action == "get_account_hierarchy":
                return await self._action_account_hierarchy(client, params)
            report_type = _report_type_for_action(action)
            if report_type:
                return await self._action_curated_report(client, params, report_type)
        except Exception as exc:
            logger.exception("Google Ads action failed")
            return ActionResponse.failure(str(exc)).to_response(status_code=500)
        return ActionResponse.not_supported(action).to_response(status_code=404)

    async def _action_run_gaql(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        customer_id, query, limit = _action_query_params(params)
        validation_error = validate_gaql_for_action(query)
        if validation_error:
            return ActionResponse.failure(validation_error).to_response(status_code=400)
        rows = await client.run_gaql(customer_id, query, limit=limit)
        return ActionResponse.success(
            {"rows": rows, "row_count": len(rows)}
        ).to_response()

    async def _action_export_csv(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        customer_id, query, limit = _action_query_params(
            params, default_limit=DEFAULT_EXPORT_LIMIT, max_limit=MAX_CSV_ROWS
        )
        validation_error = validate_gaql_for_action(query)
        if validation_error:
            return ActionResponse.failure(validation_error).to_response(status_code=400)
        rows = await client.run_gaql(customer_id, query, limit=limit)
        customer_context = await self._customer_report_context(client, customer_id)
        metadata = _report_metadata(
            "custom_gaql", customer_id, params, query, len(rows), limit, customer_context
        )
        csv_text = rows_to_csv(rows, metadata=metadata)
        return Response(
            content=csv_text,
            media_type="text/csv; charset=utf-8",
            headers={
                "content-disposition": 'attachment; filename="google-ads-report.csv"'
            },
        )

    async def _action_export_xlsx(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        customer_id, query, limit = _action_query_params(
            params, default_limit=DEFAULT_EXPORT_LIMIT, max_limit=MAX_XLSX_ROWS
        )
        validation_error = validate_gaql_for_action(query)
        if validation_error:
            return ActionResponse.failure(validation_error).to_response(status_code=400)
        rows = await client.run_gaql(customer_id, query, limit=limit)
        customer_context = await self._customer_report_context(client, customer_id)
        metadata = _report_metadata(
            "custom_gaql", customer_id, params, query, len(rows), limit, customer_context
        )
        content = rows_to_xlsx(rows, metadata=metadata)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "content-disposition": 'attachment; filename="google-ads-report.xlsx"'
            },
        )

    async def _action_account_summary(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        customer_id = _require_customer_id(params)
        date_range = str(params.get("date_range") or "LAST_30_DAYS")
        query = f"""
            SELECT
              campaign.id,
              campaign.name,
              campaign.status,
              campaign.advertising_channel_type,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.conversions,
              metrics.conversions_value
            FROM campaign
            WHERE segments.date DURING {date_range}
            ORDER BY metrics.cost_micros DESC
        """
        rows = await client.run_gaql(customer_id, query, limit=1000)
        return ActionResponse.success(
            {"customer_id": customer_id, "date_range": date_range, "rows": rows}
        ).to_response()

    async def _action_recommendations(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        customer_id = _require_customer_id(params)
        rows = await client.run_gaql(
            customer_id, SYNC_QUERIES["recommendation"], limit=1000
        )
        return ActionResponse.success(
            {"customer_id": customer_id, "recommendations": rows}
        ).to_response()

    async def _action_curated_report(
        self, client: GoogleAdsClient, params: dict[str, Any], report_type: str
    ) -> Response:
        customer_id = _require_customer_id(params)
        limit = _row_limit(params, DEFAULT_REPORT_LIMIT, MAX_JSON_ROWS)
        query = build_report_query(report_type, params)
        rows = await client.run_gaql(customer_id, query, limit=limit)
        customer_context = await self._customer_report_context(client, customer_id)
        metadata = _report_metadata(
            report_type, customer_id, params, query, len(rows), limit, customer_context
        )
        return ActionResponse.success({"metadata": metadata, "rows": rows}).to_response()

    async def _action_export_report(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        customer_id = _require_customer_id(params)
        report_type = str(params.get("report_type") or "").strip()
        if report_type not in REPORT_BUILDERS:
            return ActionResponse.failure(f"Unsupported report_type: {report_type}").to_response(
                status_code=400
            )
        output_format = str(params.get("format") or "csv").lower().strip()
        max_rows = MAX_XLSX_ROWS if output_format == "xlsx" else MAX_CSV_ROWS
        if output_format == "json":
            max_rows = MAX_JSON_ROWS
        limit = _row_limit(params, DEFAULT_EXPORT_LIMIT, max_rows)
        query = build_report_query(report_type, params)
        rows = await client.run_gaql(customer_id, query, limit=limit)
        customer_context = await self._customer_report_context(client, customer_id)
        metadata = _report_metadata(
            report_type, customer_id, params, query, len(rows), limit, customer_context
        )
        if output_format == "json":
            return ActionResponse.success({"metadata": metadata, "rows": rows}).to_response()
        if output_format == "xlsx":
            content = rows_to_xlsx(rows, metadata=metadata)
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "content-disposition": f'attachment; filename="google-ads-{report_type}.xlsx"'
                },
            )
        if output_format != "csv":
            return ActionResponse.failure(
                f"Unsupported export format: {output_format}"
            ).to_response(status_code=400)
        return Response(
            content=rows_to_csv(rows, metadata=metadata),
            media_type="text/csv; charset=utf-8",
            headers={
                "content-disposition": f'attachment; filename="google-ads-{report_type}.csv"'
            },
        )

    async def _customer_report_context(
        self, client: GoogleAdsClient, customer_id: str
    ) -> dict[str, Any]:
        query = """
            SELECT
              customer.id,
              customer.descriptive_name,
              customer.currency_code,
              customer.time_zone
            FROM customer
        """
        try:
            rows = await client.run_gaql(customer_id, query, limit=1)
        except GoogleAdsConnectorError as exc:
            logger.warning("Failed to fetch Google Ads customer report context: %s", exc)
            return {}
        if not rows:
            return {}
        customer = rows[0].get("customer", rows[0])
        return cast(dict[str, Any], customer if isinstance(customer, dict) else {})

    async def _action_account_hierarchy(
        self, client: GoogleAdsClient, params: dict[str, Any]
    ) -> Response:
        limit = _row_limit(params, 1000, MAX_JSON_ROWS)
        accessible_customers = await client.list_accessible_customers()
        customer_id = _optional_customer_id(params)
        hierarchy_rows: list[dict[str, Any]] = []
        if customer_id:
            query = """
                SELECT
                  customer_client.client_customer,
                  customer_client.descriptive_name,
                  customer_client.id,
                  customer_client.level,
                  customer_client.manager,
                  customer_client.status,
                  customer_client.currency_code,
                  customer_client.time_zone
                FROM customer_client
                LIMIT 10000
            """
            hierarchy_rows = await client.run_gaql(customer_id, query, limit=limit)
        return ActionResponse.success(
            {
                "accessible_customers": accessible_customers,
                "customer_id": customer_id,
                "hierarchy": hierarchy_rows,
                "row_count": len(hierarchy_rows),
                "limit": limit,
            }
        ).to_response()



def _optional_customer_id(params: dict[str, Any]) -> str | None:
    raw = params.get("customer_id")
    if raw is None:
        return None
    normalized = str(raw).replace("-", "").strip()
    return normalized or None


def _has_configured_customer_ids(
    source_config: dict[str, Any], credentials: dict[str, Any]
) -> bool:
    return any(
        source_config.get(key) or credentials.get(key)
        for key in (
            "customer_ids",
            "selected_customer_ids",
            "customer_id",
            "selected_customer_id",
        )
    )


def _require_customer_id(params: dict[str, Any]) -> str:
    customer_id = _optional_customer_id(params)
    if not customer_id:
        raise ValueError("Missing customer_id")
    return customer_id


def _row_limit(params: dict[str, Any], default_limit: int, max_limit: int) -> int:
    raw = params.get("limit")
    limit = default_limit if raw in (None, "") else int(raw)
    return min(max(limit, 1), max_limit)


def _action_query_params(
    params: dict[str, Any], default_limit: int = 1000, max_limit: int = MAX_JSON_ROWS
) -> tuple[str, str, int]:
    customer_id = _require_customer_id(params)
    query = str(params.get("query") or "").strip()
    if not query:
        raise ValueError("Missing query")
    limit = _row_limit(params, default_limit, max_limit)
    return customer_id, query, limit


def validate_gaql_for_action(query: str) -> str | None:
    lowered = " ".join(query.lower().split())
    if not lowered.startswith("select ") or " from " not in lowered:
        return "Only SELECT GAQL queries are supported"
    forbidden = [" mutate ", " insert ", " update ", " delete ", ";", "--", "/*"]
    if any(token in lowered for token in forbidden):
        return "Query contains unsupported tokens"
    resource = lowered.split(" from ", 1)[1].split()[0]
    if resource not in REPORT_RESOURCE_ALLOWLIST and resource not in SYNC_QUERIES:
        return f"Unsupported GAQL resource: {resource}"
    return None


def _report_common_schema_properties() -> dict[str, Any]:
    return {
        "date_range": {"type": "string", "default": "LAST_30_DAYS"},
        "start_date": {"type": "string", "description": "YYYY-MM-DD inclusive"},
        "end_date": {"type": "string", "description": "YYYY-MM-DD inclusive"},
        "limit": {"type": "integer", "default": DEFAULT_REPORT_LIMIT, "maximum": MAX_JSON_ROWS},
    }


def _report_action_schema(
    *, extra_properties: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "type": "object",
        "properties": {
            "customer_id": {"type": "string"},
            **_report_common_schema_properties(),
            **(extra_properties or {}),
        },
        "required": ["customer_id"],
    }


def _gaql_export_schema() -> dict[str, Any]:
    return {
        "type": "object",
        "properties": {
            "customer_id": {"type": "string"},
            "query": {"type": "string"},
            "limit": {
                "type": "integer",
                "default": DEFAULT_EXPORT_LIMIT,
                "maximum": MAX_CSV_ROWS,
                "description": (
                    f"Maximum rows to export. Server caps CSV at {MAX_CSV_ROWS} rows "
                    f"and XLSX at {MAX_XLSX_ROWS} rows."
                ),
            },
        },
        "required": ["customer_id", "query"],
    }


def _report_action_definition(
    name: str,
    description: str,
    *,
    extra_properties: dict[str, Any] | None = None,
) -> ActionDefinition:
    return ActionDefinition(
        name=name,
        description=description,
        mode="read",
        source_types=["google_ads"],
        input_schema=_report_action_schema(extra_properties=extra_properties),
    )


def _report_type_for_action(action: str) -> str | None:
    prefix = "get_"
    suffix = "_report"
    if not action.startswith(prefix) or not action.endswith(suffix):
        return None
    report_type = action[len(prefix) : -len(suffix)]
    if report_type == "ad_creative":
        return "ad_creative"
    if report_type in REPORT_BUILDERS:
        return report_type
    return None


def _date_where_clause(params: dict[str, Any]) -> str:
    start_date = str(params.get("start_date") or "").strip()
    end_date = str(params.get("end_date") or "").strip()
    if start_date and end_date:
        _validate_date_literal(start_date)
        _validate_date_literal(end_date)
        return f"segments.date BETWEEN '{start_date}' AND '{end_date}'"
    date_range = str(params.get("date_range") or "LAST_30_DAYS").strip().upper()
    if not date_range.replace("_", "").isalnum():
        raise ValueError("date_range must be a Google Ads DURING literal like LAST_30_DAYS")
    return f"segments.date DURING {date_range}"


def _validate_date_literal(value: str) -> None:
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"Invalid date literal: {value}. Expected YYYY-MM-DD") from exc


def _change_event_where_clause(params: dict[str, Any]) -> str:
    start_date = str(params.get("start_date") or "").strip()
    end_date = str(params.get("end_date") or "").strip()
    if start_date and end_date:
        _validate_date_literal(start_date)
        _validate_date_literal(end_date)
        return (
            f"change_event.change_date_time >= '{start_date} 00:00:00' "
            f"AND change_event.change_date_time <= '{end_date} 23:59:59'"
        )
    since = (datetime.now(UTC) - timedelta(days=14)).strftime("%Y-%m-%d 00:00:00")
    return f"change_event.change_date_time >= '{since}'"


def build_report_query(report_type: str, params: dict[str, Any]) -> str:
    where = _date_where_clause(params)
    if report_type == "campaign_performance":
        return f"""
            SELECT
              segments.date,
              campaign.id,
              campaign.name,
              campaign.status,
              campaign.advertising_channel_type,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.ctr,
              metrics.average_cpc,
              metrics.conversions,
              metrics.conversions_value
            FROM campaign
            WHERE {where}
            ORDER BY segments.date DESC, metrics.cost_micros DESC
        """
    if report_type == "ad_group_performance":
        return f"""
            SELECT
              segments.date,
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              ad_group.status,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.ctr,
              metrics.average_cpc,
              metrics.conversions,
              metrics.conversions_value
            FROM ad_group
            WHERE {where}
            ORDER BY segments.date DESC, metrics.cost_micros DESC
        """
    if report_type == "keyword_performance":
        return f"""
            SELECT
              segments.date,
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              ad_group_criterion.criterion_id,
              ad_group_criterion.keyword.text,
              ad_group_criterion.keyword.match_type,
              ad_group_criterion.status,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.ctr,
              metrics.average_cpc,
              metrics.conversions,
              metrics.conversions_value
            FROM keyword_view
            WHERE {where}
            ORDER BY metrics.cost_micros DESC
        """
    if report_type == "search_term":
        return f"""
            SELECT
              segments.date,
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              search_term_view.search_term,
              search_term_view.status,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.ctr,
              metrics.average_cpc,
              metrics.conversions,
              metrics.conversions_value
            FROM search_term_view
            WHERE {where}
            ORDER BY metrics.cost_micros DESC
        """
    if report_type == "ad_creative":
        return f"""
            SELECT
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              ad_group_ad.ad.id,
              ad_group_ad.ad.name,
              ad_group_ad.status,
              ad_group_ad.ad.type,
              ad_group_ad.ad.final_urls,
              ad_group_ad.ad.responsive_search_ad.headlines,
              ad_group_ad.ad.responsive_search_ad.descriptions,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.ctr,
              metrics.conversions,
              metrics.conversions_value
            FROM ad_group_ad
            WHERE {where}
            ORDER BY metrics.impressions DESC
        """
    if report_type == "asset_performance":
        return f"""
            SELECT
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              ad_group_ad_asset_view.field_type,
              ad_group_ad_asset_view.performance_label,
              asset.id,
              asset.name,
              asset.type,
              asset.text_asset.text,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.conversions,
              metrics.conversions_value
            FROM ad_group_ad_asset_view
            WHERE {where}
            ORDER BY metrics.impressions DESC
        """
    if report_type == "landing_page":
        return f"""
            SELECT
              segments.date,
              expanded_landing_page_view.expanded_final_url,
              metrics.impressions,
              metrics.clicks,
              metrics.cost_micros,
              metrics.ctr,
              metrics.average_cpc,
              metrics.conversions,
              metrics.conversions_value
            FROM expanded_landing_page_view
            WHERE {where}
            ORDER BY metrics.cost_micros DESC
        """
    if report_type == "budget_pacing":
        return f"""
            SELECT
              campaign_budget.id,
              campaign_budget.name,
              campaign_budget.amount_micros,
              campaign_budget.status,
              campaign.id,
              campaign.name,
              campaign.status,
              campaign.serving_status,
              metrics.cost_micros,
              metrics.impressions,
              metrics.clicks,
              metrics.conversions,
              metrics.conversions_value
            FROM campaign
            WHERE {where}
            ORDER BY metrics.cost_micros DESC
        """
    if report_type == "conversion_performance":
        return f"""
            SELECT
              segments.date,
              segments.conversion_action,
              segments.conversion_action_name,
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              metrics.conversions,
              metrics.conversions_value,
              metrics.cost_micros,
              metrics.cost_per_conversion
            FROM ad_group
            WHERE {where}
            ORDER BY metrics.conversions_value DESC
        """
    if report_type == "change_history":
        change_where = _change_event_where_clause(params)
        return f"""
            SELECT
              change_event.change_date_time,
              change_event.change_resource_type,
              change_event.changed_fields,
              change_event.client_type,
              change_event.resource_change_operation,
              change_event.resource_name,
              change_event.user_email
            FROM change_event
            WHERE {change_where}
            ORDER BY change_event.change_date_time DESC
        """
    if report_type == "policy_diagnostics":
        return """
            SELECT
              campaign.id,
              campaign.name,
              ad_group.id,
              ad_group.name,
              ad_group_ad.ad.id,
              ad_group_ad.status,
              ad_group_ad.policy_summary.approval_status,
              ad_group_ad.policy_summary.review_status,
              ad_group_ad.policy_summary.policy_topic_entries
            FROM ad_group_ad
            ORDER BY campaign.name, ad_group.name
        """
    if report_type == "segmented_performance":
        return _build_segmented_query(params, where)
    raise ValueError(f"Unsupported report_type: {report_type}")


def _build_segmented_query(params: dict[str, Any], where: str) -> str:
    resource = str(params.get("resource") or "campaign").strip()
    if resource not in SEGMENTED_REPORT_RESOURCES:
        raise ValueError(f"Unsupported segmented report resource: {resource}")
    dimensions = _string_list_param(params.get("dimensions")) or ["segments.date"]
    metrics = _string_list_param(params.get("metrics")) or [
        "metrics.impressions",
        "metrics.clicks",
        "metrics.cost_micros",
        "metrics.conversions",
        "metrics.conversions_value",
    ]
    unsupported_dimensions = [d for d in dimensions if d not in SEGMENT_FIELDS]
    unsupported_metrics = [m for m in metrics if m not in METRIC_FIELDS]
    if unsupported_dimensions:
        raise ValueError(f"Unsupported dimensions: {', '.join(unsupported_dimensions)}")
    if unsupported_metrics:
        raise ValueError(f"Unsupported metrics: {', '.join(unsupported_metrics)}")
    base_fields = {
        "customer": ["customer.id", "customer.descriptive_name"],
        "campaign": ["campaign.id", "campaign.name", "campaign.status"],
        "ad_group": ["campaign.id", "campaign.name", "ad_group.id", "ad_group.name"],
        "ad_group_ad": [
            "campaign.id",
            "campaign.name",
            "ad_group.id",
            "ad_group.name",
            "ad_group_ad.ad.id",
            "ad_group_ad.status",
        ],
        "keyword_view": [
            "campaign.id",
            "campaign.name",
            "ad_group.id",
            "ad_group.name",
            "ad_group_criterion.criterion_id",
            "ad_group_criterion.keyword.text",
        ],
    }[resource]
    fields = [*dimensions, *base_fields, *metrics]
    return f"""
        SELECT
          {', '.join(fields)}
        FROM {resource}
        WHERE {where}
        ORDER BY metrics.cost_micros DESC
    """


def _string_list_param(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [v.strip() for v in value.split(",") if v.strip()]
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return []


def _report_metadata(
    report_type: str,
    customer_id: str,
    params: dict[str, Any],
    query: str,
    row_count: int,
    limit: int,
    customer_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    customer_context = customer_context or {}
    return {
        "report_type": report_type,
        "customer_id": customer_id,
        "customer_descriptive_name": customer_context.get("descriptive_name"),
        "customer_currency_code": customer_context.get("currency_code"),
        "customer_time_zone": customer_context.get("time_zone"),
        "date_range": params.get("date_range"),
        "start_date": params.get("start_date"),
        "end_date": params.get("end_date"),
        "generated_at": datetime.now(UTC).isoformat(),
        "row_count": row_count,
        "limit": limit,
        "limit_reached": row_count >= limit,
        "cost_units": "metrics.cost_micros / 1,000,000",
        "attribution_note": (
            "Google Ads conversion metrics depend on account conversion settings "
            "and attribution windows and may change retroactively."
        ),
        "filter_note": "Reconcile with Google Ads UI using the same date range and filters.",
        "gaql": " ".join(query.split()),
    }


def rows_to_csv(rows: list[dict[str, Any]], metadata: dict[str, Any] | None = None) -> str:
    flattened = _flatten_report_rows(rows)
    fieldnames = sorted({key for row in flattened for key in row}) or ["result"]
    schema = _infer_schema(flattened, fieldnames)
    output = io.StringIO()
    if metadata:
        output.write("# Google Ads report metadata\n")
        for key, value in metadata.items():
            output.write(f"# {key}: {value}\n")
        output.write("# columns: " + ", ".join(fieldnames) + "\n")
        output.write("\n")
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in flattened:
        writer.writerow(row)
    if metadata:
        output.write("\n# Schema\n")
        for column in fieldnames:
            output.write(f"# {column}: {schema[column]}\n")
    return output.getvalue()


def rows_to_xlsx(rows: list[dict[str, Any]], metadata: dict[str, Any] | None = None) -> bytes:
    from openpyxl import Workbook  # type: ignore[import-untyped]

    flattened = _flatten_report_rows(rows)
    fieldnames = sorted({key for row in flattened for key in row}) or ["result"]
    schema = _infer_schema(flattened, fieldnames)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    ws.append(fieldnames)
    for row in flattened:
        ws.append([row.get(name) for name in fieldnames])
    if metadata:
        meta = wb.create_sheet("Metadata")
        meta.append(["key", "value"])
        for key, value in metadata.items():
            meta.append([key, str(value)])
    schema_ws = wb.create_sheet("Schema")
    schema_ws.append(["column", "inferred_type"])
    for column in fieldnames:
        schema_ws.append([column, schema[column]])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _flatten_report_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_with_derived_metrics(_flatten(row)) for row in rows]


def _with_derived_metrics(row: dict[str, Any]) -> dict[str, Any]:
    out = dict(row)
    cost_micros = _as_float(row.get("metrics.cost_micros"))
    impressions = _as_float(row.get("metrics.impressions"))
    clicks = _as_float(row.get("metrics.clicks"))
    conversions = _as_float(row.get("metrics.conversions"))
    conversion_value = _as_float(row.get("metrics.conversions_value"))

    if cost_micros is not None:
        out["derived.cost"] = cost_micros / 1_000_000
    cost = _as_float(out.get("derived.cost"))
    if clicks and impressions:
        out["derived.ctr"] = clicks / impressions
    if conversions and clicks:
        out["derived.conversion_rate"] = conversions / clicks
    if conversions and cost is not None:
        out["derived.cost_per_conversion"] = cost / conversions
    if conversion_value is not None and cost:
        out["derived.roas"] = conversion_value / cost
    return out


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _infer_schema(rows: list[dict[str, Any]], fieldnames: list[str]) -> dict[str, str]:
    return {field: _infer_column_type([row.get(field) for row in rows]) for field in fieldnames}


def _infer_column_type(values: list[Any]) -> str:
    present = [value for value in values if value not in (None, "")]
    if not present:
        return "empty"
    if all(isinstance(value, bool) for value in present):
        return "boolean"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in present):
        return "integer"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present):
        return "number"
    return "text"


def _flatten(value: Any, prefix: str = "") -> dict[str, Any]:
    if isinstance(value, dict):
        out: dict[str, Any] = {}
        for key, item in value.items():
            path = f"{prefix}.{key}" if prefix else key
            out.update(_flatten(item, path))
        return out
    if isinstance(value, list):
        return {prefix: ", ".join(str(v) for v in value)}
    return {prefix: value}
